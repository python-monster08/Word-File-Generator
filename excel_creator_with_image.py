import PyPDF2
import fitz
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import io

# Function to extract text from PDF
def extract_text(pdf_path):
    text = ""
    with open(pdf_path, "rb") as file:
        reader = PyPDF2.PdfReader(file)
        for page_num in range(len(reader.pages)):
            page = reader.pages[page_num]
            text += page.extract_text()
    return text

# Function to extract images from PDF
def extract_images(pdf_path):
    images = []
    pdf_document = fitz.open(pdf_path)
    for page_num in range(len(pdf_document)):
        page = pdf_document.load_page(page_num)
        image_list = page.get_images(full=True)
        for img_index, img in enumerate(image_list):
            xref = img[0]
            base_image = pdf_document.extract_image(xref)
            image_bytes = base_image["image"]
            images.append(image_bytes)
    return images

# Create a new Excel workbook and worksheet
wb = Workbook()
ws = wb.active

# Extract text and images from PDF
pdf_path = 'Full Disha 26 2020 Years Prelims Solved Papers Ancient_History.pdf'
pdf_text = extract_text(pdf_path)
pdf_images = extract_images(pdf_path)

# Write text to Excel
ws['A1'] = pdf_text

# Insert images into Excel
for index, image_bytes in enumerate(pdf_images):
    try:
        img = Image(io.BytesIO(image_bytes))
        # Calculate the position to insert the image
        row = len(pdf_text.split('\n')) + 2  # Place images below text
        img.anchor = f'{get_column_letter(1)}{row + index * 20}'  # Adjust row based on image size
        ws.add_image(img)
    except Exception as e:
        print(f"Error processing image {index + 1}: {e}")

# Save the Excel file
wb.save('pdf_content.xlsx')
